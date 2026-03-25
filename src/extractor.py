#!/usr/bin/env python3

import json
import logging
from datetime import timedelta
from pathlib import Path
from typing import Dict
from typing import List
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

from report import summarize_excel_dataframe


BROKEN_BEAT_GENRES = [
    "Breakbeat",
    "Breakcore",
    "Breaks",
    "Progressive Breaks",
    "Broken Beat",
    "Drum n Bass",
    "Jungle",
    "Halftime",
    "Juke",
    "UK Garage",
    "Speed Garage",
    "Bassline",
    "Electro",
]


def find_json_files(directory: Path) -> List[Path]:
    if not directory.is_dir():
        return []
    files = [item for item in directory.rglob("*.json") if item.is_file()]
    files.sort()
    return files


def format_list_with_comma(items: List[str]) -> str:
    if not items:
        return ""
    unique_items: List[str] = []
    seen = set()
    for item in items:
        if item not in seen:
            unique_items.append(item)
            seen.add(item)
    return ", ".join(unique_items)


def format_confidences(confidences: List[float]) -> str:
    if not confidences:
        return ""
    return ", ".join(str(confidence) for confidence in confidences)


def check_broken_beat(genres: List[str]) -> bool:
    return any(genre in BROKEN_BEAT_GENRES for genre in genres)


def _duration_text_to_timedelta(value: str) -> timedelta:
    text = str(value or "").strip()
    if not text:
        return timedelta(0)
    try:
        parts = text.split(":")
        if len(parts) != 3:
            return timedelta(0)
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2])
        return timedelta(hours=hours, minutes=minutes, seconds=seconds)
    except Exception:
        return timedelta(0)


def process_json_file(
    json_path: Path,
) -> Optional[Dict[str, object]]:
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
        genres_data = data.get("genres", {})
        labels = genres_data.get("labels", [])
        confidences = genres_data.get("confidences", [])
        model_key = str(genres_data.get("model", "")).strip()
        file_data = data.get("file", {})
        duration_data = file_data.get("duration", {})

        file_path_value = str(file_data.get("path", "")).strip()
        if not file_path_value:
            file_path_value = str(json_path)

        analysis_status = (
            "analysis_error"
            if str(data.get("error", "")).strip()
            else "analysis_success"
        )

        return {
            "status": analysis_status,
            "path": file_path_value,
            "name": file_data.get("name", ""),
            "duration": _duration_text_to_timedelta(str(duration_data.get("time", ""))),
            "genres": format_list_with_comma(labels),
            "is_broken_beat": bool(check_broken_beat(labels)),
            "model_key": model_key,
            "confidences": format_confidences(confidences),
        }
    except Exception as exc:
        logging.error("Error processing %s: %s", json_path, exc)
        return None


def _load_existing_excel(excel_path: Path) -> pd.DataFrame:
    if not excel_path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(excel_path, engine="openpyxl")
    except Exception as exc:
        logging.error("Error reading existing Excel file %s: %s", excel_path, exc)
        return pd.DataFrame()


def _filter_new_records(
    new_data: List[Dict[str, object]],
    existing_df: pd.DataFrame,
) -> List[Dict[str, object]]:
    if existing_df.empty:
        return new_data

    existing_paths = (
        set(existing_df["path"].tolist()) if "path" in existing_df.columns else set()
    )
    return [record for record in new_data if record["path"] not in existing_paths]


def auto_adjust_excel_columns(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    header_to_column: Dict[str, str] = {}
    for column_cells in worksheet.columns:
        header_value = column_cells[0].value
        if isinstance(header_value, str) and header_value.strip():
            header_to_column[header_value.strip()] = column_cells[0].column_letter

    for column_cells in worksheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        worksheet.column_dimensions[column_letter].width = max_len + 2

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    duration_column = header_to_column.get("duration")
    if duration_column:
        for row_index in range(2, worksheet.max_row + 1):
            worksheet[f"{duration_column}{row_index}"].number_format = "[h]:mm:ss"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(excel_path)


def create_excel_report(
    json_dir: Path,
    output_excel: Path,
) -> Dict[str, object]:
    json_files = find_json_files(json_dir)
    if not json_files:
        raise RuntimeError(f"No JSON files found in {json_dir}")

    rows: List[Dict[str, object]] = []
    for json_file in json_files:
        row_data = process_json_file(json_file)
        if row_data:
            rows.append(row_data)

    if not rows:
        raise RuntimeError("No valid records extracted from JSON files")

    existing_df = _load_existing_excel(output_excel)
    new_rows = _filter_new_records(rows, existing_df)

    if not new_rows:
        logging.info("Excel is up to date, no new records")
        library_summary = summarize_excel_dataframe(existing_df)
        return {
            "json_files": len(json_files),
            "rows_added": 0,
            "rows_total": len(existing_df),
            "library_summary": library_summary,
        }

    new_df = pd.DataFrame(new_rows)
    final_df = (
        pd.concat([existing_df, new_df], ignore_index=True)
        if not existing_df.empty
        else new_df
    )

    ordered_columns = [
        "status",
        "path",
        "name",
        "duration",
        "genres",
        "is_broken_beat",
        "model_key",
        "confidences",
    ]
    final_df = final_df.reindex(columns=ordered_columns)
    final_df["duration"] = pd.to_timedelta(
        final_df["duration"], errors="coerce"
    ).fillna(pd.Timedelta(0))
    final_df["is_broken_beat"] = final_df["is_broken_beat"].fillna(False).astype(bool)

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    final_df.to_excel(output_excel, index=False, engine="openpyxl")
    auto_adjust_excel_columns(output_excel)

    library_summary = summarize_excel_dataframe(final_df)
    return {
        "json_files": len(json_files),
        "rows_added": len(new_rows),
        "rows_total": len(final_df),
        "library_summary": library_summary,
    }
